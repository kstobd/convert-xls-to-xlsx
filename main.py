from io import TextIOWrapper
from os import ftruncate

from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

import xlrd
import openpyxl


class PreparingSchudlel:

    @staticmethod
    def Cvt_xls_to_xlsx(*args, **kw):
        book_xls = xlrd.open_workbook(*args, formatting_info=True, ragged_rows=True, **kw)      #open xls book
        book_xlsx = openpyxl.workbook.Workbook()    #create xlsx book

        sheet_names = book_xls.sheet_names()    #create name for xls book
        for sheet_index in range(len(sheet_names)):
            sheet_xls = book_xls.sheet_by_name(sheet_names[sheet_index])
            #choose sheet index
            if sheet_index == 0:
                sheet_xlsx = book_xlsx.active
                sheet_xlsx.title = sheet_names[sheet_index]
            else:
                sheet_xlsx = book_xlsx.create_sheet(title=sheet_names[sheet_index])
            #make format
            for crange in sheet_xls.merged_cells:
                rlo, rhi, clo, chi = crange

                sheet_xlsx.merge_cells(
                    start_row=rlo + 1, end_row=rhi,
                    start_column=clo + 1, end_column=chi,
                )

                #fill format book by value
                for row in range(0, sheet_xls.nrows):
                    for col in range(0, sheet_xls.ncols):
                        try:
                            sheet_xlsx.cell(row = row+1 , column = col+1).value = sheet_xls.cell_value(row, col)
                        except BaseException:
                            pass

        book_xlsx.save(filename = fname[:-3] + 'xlsx')


    @staticmethod
    def Unmerged(xlsxFile):
        wb = load_workbook(filename = xlsxFile)

        for st_name in wb.sheetnames:
            st = wb[st_name]
            mcr_coord_list = [mcr.coord for mcr in st.merged_cells.ranges]
            
            for mcr in mcr_coord_list:
                min_col, min_row, max_col, max_row = range_boundaries(mcr)
                top_left_cell_value = st.cell(row=min_row, column=min_col).value
                st.unmerge_cells(mcr)
                for row in st.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                    for cell in row:
                        cell.value = top_left_cell_value

        wb.save(fname[:-4] + '_unmegred.xlsx')




# test command

# fname = "Расписание занятий ф-т ИТ -  02.03.02, 09.03.01, 09.03.02 - 1, 2 курс - семестр 1,3 -2021-22 уч г (0916).xls"
# PreparingSchudlel().Cvt_xls_to_xlsx(fname)
# PreparingSchudlel().Unmerged("Расписание занятий ф-т ИТ -  02.03.02, 09.03.01, 09.03.02 - 1, 2 курс - семестр 1,3 -2021-22 уч г (0916).xlsx")